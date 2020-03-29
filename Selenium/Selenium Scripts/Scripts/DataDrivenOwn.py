from selenium import webdriver
import openpyxl
import time


path="C://Users/anaso/PycharmProjects/Selenium/Selenium Scripts/Data/Login2.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.get_sheet_by_name("Sheet1")
rows=sheet.max_row
columns=sheet.max_column
print(rows,columns)

driver=webdriver.Chrome("C:/Users/anaso/PycharmProjects/Selenium/Selenium Scripts/Drivers/chromedriver.exe")
driver.maximize_window()
time.sleep(2)
driver.get("https://opensource-demo.orangehrmlive.com/index.php/auth/login")

for i in range(2,rows+1):
    UserName=sheet.cell(i,1).value
    PassWord= sheet.cell(i, 2).value
    driver.find_element_by_id("txtUsername").send_keys(UserName)
    driver.find_element_by_id("txtPassword").send_keys(PassWord)
    time.sleep(1)
    driver.find_element_by_id("btnLogin").click()
    time.sleep(2)
    curl=driver.current_url
    if(curl=="https://opensource-demo.orangehrmlive.com/index.php/dashboard"):
        print("Test Passed")
        sheet.cell(i,3).value="Test Passed"
        driver.find_element_by_link_text("Welcome Admin").click()
        time.sleep(2)
        driver.find_element_by_css_selector("#welcome-menu > ul > li:nth-child(2) > a").click()

    else:
        print("Test Failed")
        sheet.cell(i, 3).value = "Test Failed"
    workbook.save(path)




