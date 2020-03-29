import openpyxl

path="C://Users/anaso/PycharmProjects/Selenium/Selenium Scripts/Data/Data.xlsx"
WBook=openpyxl.load_workbook(path)

#sheet=WBook.get_sheet_by_name("Sheet1")

#since i have only 1 work sheet
sheet=WBook.active

rows=sheet.max_row
column=sheet.max_column

print(rows,column)


for i in range(1,rows+1):
    for j in range(1,column+1):
        val=sheet.cell(i, j).value
        print("{:10}".format(val),end=" ")
    print()