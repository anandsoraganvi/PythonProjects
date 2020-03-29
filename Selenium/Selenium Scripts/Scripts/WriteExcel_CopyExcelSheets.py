import openpyxl

path1="C://Users/anaso/PycharmProjects/Selenium/Selenium Scripts/Data/Data_Create.xlsx"
WBook1=openpyxl.load_workbook(path1)

path2="C://Users/anaso/PycharmProjects/Selenium/Selenium Scripts/Data/Data.xlsx"
WBook2=openpyxl.load_workbook(path2)


sheet1=WBook1.active
sheet2=WBook2.active

rows=sheet2.max_row
column=sheet2.max_column

for i in range(1,rows+1):
    for j in range(1,column+1):
        val=sheet2.cell(i, j).value #reading from one excel sheet
        sheet1.cell(i,j).value=val  #writing to another excel sheet


WBook1.save(path1)
