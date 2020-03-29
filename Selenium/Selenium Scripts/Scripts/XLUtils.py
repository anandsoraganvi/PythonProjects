import openpyxl

def getRowcount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    rows=sheet.max_row
    return rows

def getColumncount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    columns=sheet.max_column
    return columns

def readData(file,sheetname,rownum,columnnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.cell(rownum,columnnum).value

def writeData(file,sheetname,rownum,columnnum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    sheet.cell(rownum,columnnum).value=data
    workbook.save(file)


